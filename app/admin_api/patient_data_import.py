from werkzeug.datastructures import FileStorage
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from web_errors import WebError
from dataclasses import dataclass
from typing import Iterable
from clinics.data_access import get_most_common_clinic
from patients.data_access import patient_from_key_data, add_patient
from patients.patient import Patient
from visits.data_access import first_visit_by_patient_and_date, add_visit
from visits.visit import Visit
from events.data_access import clear_all_events, add_event
from events.event import Event
import uuid
from language_strings.language_string import LanguageString
from util import as_string
from datetime import date, timedelta, datetime
import itertools
import json
from config import DEFAULT_PROVIDER_ID_FOR_IMPORT
import pandas as pd
import dateutil

COLUMNS = [
    'camp',
    'visit_date',
    'visit_type',
    'first_name',
    'surname',
    'date_of_birth',
    'age',
    'gender',
    'hometown',
    'home_country',
    'phone',
    'doctor',
    # medical hx
    'allergies',
    'surgery_hx',
    'chronic_conditions',
    'current_medications',
    'vaccinations',
    # complaint
    'complaint',
    # vitals
    'heart_rate',
    'blood_pressure',
    'sats',
    'temp',
    'respiratory_rate',
    'weight',
    'blood_glucose',
    # examination
    'examination',
    'general_observations',
    'diagnosis',
    'treatment',
    'covid_19',
    'referral',
    # medicines_1
    'medication_1',
    'type_1',
    'dosage_1',
    'days_1',
    # medicines_2
    'medication_2',
    'type_2',
    'dosage_2',
    'days_2',
    # medicines_3
    'medication_3',
    'type_3',
    'dosage_3',
    'days_3',
    # medicines_4
    'medication_4',
    'type_4',
    'dosage_4',
    'days_4',
    # medicines_5
    'medication_5',
    'type_5',
    'dosage_5',
    'days_5',
    # physiotherapy
    'previous_treatment',
    'complaint_p',
    'findings',
    'treatment_plan',
    'treatment_session',
    'recommendations',
    'referral',
    'dental_treatment',
    'notes',
    'covid_19_result',
    # deprecated
    'examination_d',
    'medical_hx_d',
    'treatment_d',
    'diagnosis_d',
    'medicine_dispensed_d',
    'prescriptions_d',
    'allergies_d',
]

@dataclass
class PatientDataRow:
    camp: str = None
    visit_date: str = None
    visit_type: str = None
    first_name: str = None
    surname: str = None
    date_of_birth: str = None
    age: str = None
    gender: str = None
    hometown: str = None
    home_country: str = None
    phone: str = None
    doctor: str = None
    allergies: str = None
    surgery_hx: str = None
    chronic_conditions: str = None
    current_medications: str = None
    vaccinations: str = None
    complaint: str = None
    heart_rate: str = None
    blood_pressure: str = None
    sats: str = None
    temp: str = None
    respiratory_rate: str = None
    weight: str = None
    blood_glucose: str = None
    examination: str = None
    general_observations: str = None
    diagnosis: str = None
    treatment: str = None
    covid_19: str = None
    referral: str = None
    medication_1: str = None
    type_1: str = None
    dosage_1: str = None
    days_1: str = None
    medication_2: str = None
    type_2: str = None
    dosage_2: str = None
    days_2: str = None
    medication_3: str = None
    type_3: str = None
    dosage_3: str = None
    days_3: str = None
    medication_4: str = None
    type_4: str = None
    dosage_4: str = None
    days_4: str = None
    medication_5: str = None
    type_5: str = None
    dosage_5: str = None
    days_5: str = None
    previous_treatment: str = None
    complaint_p: str = None
    findings: str = None
    treatment_plan: str = None
    treatment_session: str = None
    recommendations: str = None
    referral: str = None
    dental_treatment: str = None
    notes: str = None
    covid_19_result: str = None
    examination_d: str = None
    medical_hx_d: str = None
    treatment_d: str = None
    diagnosis_d: str = None
    medicine_dispensed_d: str = None
    prescriptions_d: str = None
    allergies_d: str = None


# COLUMN_TYPES = [str, None, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, float, str,
#                 float, float, float, float, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str,
#                 str, str]


# class PatientDataImporter:
#     def __init__(self, data_file: FileStorage):
#         self.data_filename = self._write_file_to_tempfile(data_file)

#     def run(self):
#         all_rows = [self._parse_row(row) for row in self.iter_data_rows()]
#         print('Creating patients...')
#         self._create_patients(all_rows)
#         print('Creating visits...')
#         self._create_visits(all_rows)

#     def _parse_row(self, row):
#         if len(row) != 41:
#             raise WebError('All data rows must have exactly 41 data points.', 400)
#         values = [self._parse_cell(value, formatter) for value, formatter in zip(row, COLUMN_TYPES)]
#         return PatientDataRow(**dict(zip(COLUMNS, values)))

#     def _parse_cell(self, cell, formatter):
#         if cell == 'Nil' or cell is None:
#             return None
#         if formatter is None:
#             return cell
#         return formatter(cell)

#     @staticmethod
#     def _write_file_to_tempfile(data_file: FileStorage):
#         handle = NamedTemporaryFile('wb', delete=False, suffix='.xlsx')
#         data_file.save(handle)
#         handle.close()
#         print('Upload written to', handle.name)
#         return handle.name

#     def iter_data_rows(self):
#         wb = load_workbook(self.data_filename)
#         ws = wb.active
#         for idx, row in enumerate(ws.iter_rows(min_row=3, max_col=41, values_only=True)):
#             if all(x is None for x in row):
#                 continue
#             yield row

#     def _create_patients(self, rows: Iterable[PatientDataRow]):
#         for patient_data in set(map(lambda r: (r.first_name, r.surname, r.gender, r.home_country, r.age), rows)):
#             first_name, surname, gender, home_country, age = patient_data
#             if not patient_from_key_data(first_name, surname, home_country, self._parse_sex(gender)):
#                 self._create_patient(first_name, surname, home_country, gender, age)

#     def _create_patient(self, given_name, surname, home_country, sex, age):
#         given_name_ls = LanguageString(id=str(uuid.uuid4()), content_by_language={'en': given_name})
#         surname_ls = LanguageString(id=str(uuid.uuid4()), content_by_language={'en': surname})
#         inferred_dob = self._infer_dob(age)
#         patient = Patient(
#             id=str(uuid.uuid4()),
#             edited_at=datetime.now(),
#             given_name=given_name_ls,
#             surname=surname_ls,
#             date_of_birth=inferred_dob,
#             sex=self._parse_sex(sex),
#             country=LanguageString(id=str(uuid.uuid4()), content_by_language={'en': home_country}),
#             phone=None,
#             hometown=None
#         )
#         add_patient(patient)

#     @staticmethod
#     def _parse_sex(sex_str):
#         if sex_str is None:
#             return None
#         elif 'm' in sex_str.lower():
#             return 'M'
#         elif 'f' in sex_str.lower():
#             return 'F'
#         else:
#             return None

#     def _infer_dob(self, age_string):
#         try:
#             int_prefix = int(''.join(itertools.takewhile(str.isnumeric, age_string)))
#             today = date.today()
#             if 'months' in age_string:
#                 return today - timedelta(days=30 * int_prefix)
#             elif 'weeks' in age_string:
#                 return today - timedelta(weeks=int_prefix)
#             elif 'days' in age_string:
#                 return today - timedelta(days=int_prefix)
#             else:
#                 # Assume years if no unit is specified
#                 return today - timedelta(days=365 * int_prefix)
#         except (ValueError, TypeError):
#             return date(1900, 1, 1)

#     @staticmethod
#     def _parse_date(date_str):
#         if isinstance(date_str, date) or isinstance(date_str, datetime):
#             return date_str
#         try:
#             dt = pd.to_datetime(date_str, dayfirst=True).to_pydatetime()
#             return date(year=dt.year, month=dt.month, day=dt.day)
#         except dateutil.parser._parser.ParserError:
#             return None

#     def _create_visits(self, rows: Iterable[PatientDataRow]):
#         for row in rows:
#             patient_id = patient_from_key_data(row.first_name, row.surname, row.home_country, self._parse_sex(row.gender))
#             if not patient_id:
#                 print('Warning: unknown patient; skipping.')
#                 continue
#             visit_date = self._parse_date(row.visit_date)
#             visit_id, visit_timestamp = first_visit_by_patient_and_date(patient_id, visit_date)

#             # TODO: The data import format does not currently specify a clinic. Since
#             # current Hikma instances are single clinic anyway, just get the most common
#             # clinic (in case there is a demo one with few if any visits) and use that.
#             clinic_id = get_most_common_clinic()

#             # TODO: The data import format does not currently specify a provider in a format
#             # that we can use. So for now, use a per-instance default provider that is set via
#             # environment variable.
#             provider_id = DEFAULT_PROVIDER_ID_FOR_IMPORT

#             if visit_id is None:
#                 visit_id = str(uuid.uuid4())
#                 visit_timestamp = datetime.combine(visit_date, datetime.min.time())
#                 visit = Visit(
#                     id=visit_id,
#                     patient_id=patient_id,
#                     edited_at=datetime.now(),
#                     clinic_id=clinic_id,
#                     provider_id=provider_id,
#                     check_in_timestamp=visit_timestamp
#                 )
#                 add_visit(visit)

#                 # Until we implement full deletion, only add visit the first time it is seen.
#                 self._update_events(patient_id, visit_id, visit_timestamp, row)

#     def _update_events(self, patient_id: str, visit_id: str, visit_timestamp: datetime, row: PatientDataRow):
#         # TODO: This will need to be replaced with a mode of deletion that persists through synchronization.
#         # clear_all_events(visit_id)
#         if row.allergies:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Allergies', row.allergies)
#         if any([row.dispensed_medicine_1, row.dispensed_medicine_2,
#                 row.dispensed_medicine_3, row.dispensed_medicine_4]):
#             self._add_dispensed_medicine_event(patient_id, visit_id, visit_timestamp, row)
#         if row.presenting_complaint:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Complaint', row.presenting_complaint)
#         if any([row.heart_rate, row.blood_pressure, row.o2_sats,
#                 row.respiratory_rate, row.temperature, row.blood_glucose]):
#             self._add_vitals_event(patient_id, visit_id, visit_timestamp, row)
#         if row.examination:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Examination', row.examination)
#         if row.diagnosis:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Diagnosis', row.diagnosis)
#         if row.treatment:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Treatment', row.treatment)
#         if row.prescription:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Prescriptions', row.prescription)
#         if row.notes:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Notes', row.notes)
#         if row.camp:
#             self._add_text_event(patient_id, visit_id, visit_timestamp, 'Camp', row.camp)

#     def _add_text_event(self, patient_id: str, visit_id: str, visit_timestamp: datetime,
#                         event_type: str, event_metadata: str):
#         event = Event(
#             id=str(uuid.uuid4()),
#             patient_id=patient_id,
#             visit_id=visit_id,
#             event_type=event_type,
#             event_timestamp=visit_timestamp,
#             event_metadata=event_metadata,
#             edited_at=datetime.now(),
#         )
#         add_event(event)

#     def _add_dispensed_medicine_event(self,  patient_id: str, visit_id: str, visit_timestamp: datetime, row: PatientDataRow):
#         data = [
#             (row.dispensed_medicine_1, row.dispensed_medicine_quantity_1),
#             (row.dispensed_medicine_2, row.dispensed_medicine_quantity_2),
#             (row.dispensed_medicine_3, row.dispensed_medicine_quantity_3),
#             (row.dispensed_medicine_4, row.dispensed_medicine_quantity_4),
#         ]
#         content = '\n'.join([': '.join(r) for r in data if all(r)])
#         event = Event(
#             id=str(uuid.uuid4()),
#             patient_id=patient_id,
#             visit_id=visit_id,
#             event_type='Medicine Dispensed',
#             event_timestamp=visit_timestamp,
#             event_metadata=content,
#             edited_at=datetime.now(),
#         )
#         add_event(event)

#     def _add_vitals_event(self,  patient_id: str, visit_id: str, visit_timestamp: datetime, row: PatientDataRow):
#         try:
#             diastolic, systolic = row.blood_pressure.split('/')
#         except (ValueError, AttributeError):
#             diastolic = None
#             systolic = None

#         data = {
#             'heartRate': as_string(row.heart_rate),
#             'systolic': as_string(systolic),
#             'diastolic': as_string(diastolic),
#             'sats': as_string(row.o2_sats),
#             'temp': as_string(row.temperature),
#             'respiratoryRate': as_string(row.respiratory_rate),
#             'bloodGlucose': as_string(row.blood_glucose)
#         }

#         event = Event(
#             id=str(uuid.uuid4()),
#             patient_id=patient_id,
#             visit_id=visit_id,
#             event_type='Vitals',
#             event_timestamp=visit_timestamp,
#             event_metadata=json.dumps(data),
#             edited_at=datetime.now(),
#         )
#         add_event(event)
