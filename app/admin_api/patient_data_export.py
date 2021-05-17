from admin_api.patient_data_import import PatientDataRow, COLUMNS
from visits.data_access import all_visits
from openpyxl import load_workbook
from events.data_access import events_by_visit, camp_by_patient
from patients.data_access import patient_from_id
from events.event_export import write_vitals_event, write_medical_hx_event, write_examination_event, write_med1_event, write_med2_event, write_med3_event, write_physiotherapy_event, write_covid_19_event
from datetime import datetime, timedelta
from tempfile import NamedTemporaryFile
import json
from google.cloud import storage
from config import EXPORTS_STORAGE_BUCKET


def most_recent_export():
    storage_client = storage.Client()
    blobs = storage_client.list_blobs(EXPORTS_STORAGE_BUCKET)
    most_recent = max(blobs, key=lambda b: b.name)
    output = NamedTemporaryFile('wb', suffix='.xlsx', delete=False)
    output.close()
    most_recent.download_to_filename(output.name)
    return output.name


class PatientDataExporter:
    def __init__(self):
        self.rows = []

    def run(self):
        workbook = load_workbook('data/base_export.xlsx')
        worksheet = workbook.get_sheet_by_name('Sheet1')
        for i, row in enumerate(self.iter_data_rows()):
            self.write_row(worksheet, i, row)
        output = NamedTemporaryFile('wb', suffix='.xlsx', delete=False)
        output.close()
        workbook.save(output.name)
        return output.name

    def write_row(self, worksheet, row_index, row):
        for col_index, col_key in enumerate(COLUMNS):
            value = getattr(row, col_key)
            if value is not None:
                cell = worksheet.cell(row_index + 3, col_index + 1)
                cell.value = value

    def iter_data_rows(self):
        for visit in all_visits():
            if not visit.patient_id:
                continue
            patient = patient_from_id(visit.patient_id)
            if not patient:
                continue
            row = PatientDataRow(
                visit_date=visit.check_in_timestamp,
                first_name=patient.given_name.get('en'),
                surname=patient.surname.get('en'),
                age=self.age_string_from_dob(patient.date_of_birth),
                gender=patient.sex,
                home_country=patient.country.get('en')
            )
            camp_event = camp_by_patient(visit.patient_id)
            if camp_event is not None:
                self.write_text_event(row, 'camp', camp_event.event_metadata)
            for event in events_by_visit(visit.id):
                if event.event_type == 'Visit Type':
                    self.write_text_event(row, 'visit_type', event.event_metadata)
                elif event.event_type == 'Medical History Full':
                    write_medical_hx_event(row, event)
                elif event.event_type == 'Vitals':
                    write_vitals_event(row, event)
                elif event.event_type == 'Examination Full':
                    write_examination_event(row, event)
                elif event.event_type == 'Physiotherapy':
                    write_physiotherapy_event(row, event)
                elif event.event_type == 'Medicine':
                    if row.medication_1 is None:
                        write_med1_event(row, event)
                    elif row.medication_2 is None:
                        write_med2_event(row, event)
                    elif row.medication_3 is None:
                        write_med3_event(row, event)
                elif event.event_type == 'Notes':
                    self.write_text_event(row, 'notes', event.event_metadata)
                elif event.event_type == 'Dental Treatment':
                    self.write_text_event(row, 'dental_treatment', event.event_metadata)
                elif event.event_type == 'Complaint':
                    self.write_text_event(row, 'complaint', event.event_metadata)
                elif event.event_type == 'COVID-19 Screening':
                    write_covid_19_event(row, event)

                elif event.event_type == 'Allergies':
                    self.write_text_event(row, 'allergies_d', event.event_metadata)
                elif event.event_type == 'Medicine Dispensed':
                    self.write_text_event(row, 'medicine_dispensed_d', event.event_metadata)
                elif event.event_type == 'Medical History':
                    self.write_text_event(row, 'medical_hx_d', event.event_metadata)
                elif event.event_type == 'Examination':
                    self.write_text_event(row, 'examination_d', event.event_metadata)
                elif event.event_type == 'Diagnosis':
                    self.write_text_event(row, 'diagnosis_d', event.event_metadata)
                elif event.event_type == 'Treatment':
                    self.write_text_event(row, 'treatment_d', event.event_metadata)
                elif event.event_type == 'Prescriptions':
                    self.write_text_event(row, 'prescriptions_d', event.event_metadata)
            yield row

    def write_text_event(self, row, key, text):
        setattr(row, key, text)

    def write_vitals_event(self, row: PatientDataRow, event):
        data = json.loads(event.event_metadata)
        row.heart_rate = data.get('heartRate')
        if data.get('systolic') and data.get('diastolic'):
            row.blood_pressure = f"{data.get('systolic')}/{data.get('diastolic')}"
        row.o2_sats = data.get('sats')
        row.temperature = data.get('temp')
        row.respiratory_rate = data.get('respiratoryRate')
        row.blood_glucose = data.get('bloodGlucose')

    def age_string_from_dob(self, dob):
        if dob is None:
            return 'unknown'
        age = datetime.now() - datetime(dob.year, dob.month, dob.day)
        if age < timedelta(days=365):
            return f'{(age.days // 30) + 1} months'
        return f'{(age.days // 365)} years'